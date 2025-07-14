import os
from pathlib import Path
import pandas as pd
import re
from docx import Document
from datetime import datetime
import openpyxl
import logging
import glob
import zipfile

logger = logging.getLogger(__name__)

class GradeRegistration:
    def __init__(self):
        self.homework_pattern = re.compile(r'第(\d+)次作业')
        self.grade_pattern = re.compile(r'老师评分：\s*([A-E])')
        # 默认Excel文件路径
        self.default_excel_path = Path("/Users/linyuan/jobs/22g-class-java-homework/平时成绩登记表-22计算机G1班.xlsx")
        # 设置默认仓库路径
        self.repo_path = Path(".")
        # 添加缓存字典，用于存储每个班级的学生名单
        self._student_cache = {}
    
    def find_excel_file(self, class_name: str) -> Path:
        """查找Excel文件
        
        Args:
            class_name: 班级名称
            
        Returns:
            Excel文件路径
            
        Raises:
            FileNotFoundError: 如果找不到对应的Excel文件
        """
        # 构建Excel文件名
        excel_name = f"平时成绩登记表-{class_name}.xlsx"
        excel_path = self.repo_path / excel_name
        
        # 如果文件不存在，抛出异常
        if not excel_path.exists():
            raise FileNotFoundError(f"找不到班级 {class_name} 的成绩登记表")
        
        return excel_path
    
    def get_homework_number_from_path(self, file_path: str) -> int:
        """从文件路径中提取作业次数
        
        Args:
            file_path: 文件路径
            
        Returns:
            int: 作业次数，如果无法提取则返回0
        """
        logger.info(f"尝试从路径提取作业次数: {file_path}")
        
        # 获取文件所在目录名
        dir_name = os.path.basename(os.path.dirname(file_path))
        
        # 从目录名中提取作业次数
        match = re.search(r'第([一二三四五六七八九十]+)次作业|作业([一二三四五六七八九十]+)|作业(\d+)', dir_name)
        if not match:
            logger.warning(f"无法从目录名中提取作业次数: {dir_name}")
            return 0
        
        # 获取匹配的数字（可能是中文数字或阿拉伯数字）
        number = match.group(1) or match.group(2) or match.group(3)
        
        # 如果是中文数字，转换为阿拉伯数字
        if re.match(r'[一二三四五六七八九十]+', number):
            chinese_numbers = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5,
                             '六': 6, '七': 7, '八': 8, '九': 9, '十': 10}
            return chinese_numbers.get(number, 0)
        else:
            return int(number)
    
    def _load_student_list(self, excel_path: Path) -> dict:
        """加载Excel文件中的学生名单
        
        Args:
            excel_path: Excel文件路径
            
        Returns:
            dict: 学生行号映射，key为学生姓名，value为行号
        """
        student_rows = {}
        with open(str(excel_path), 'rb') as f:
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            
            # 查找学生行
            for row_idx in range(1, ws.max_row + 1):
                cell_value = str(ws.cell(row=row_idx, column=3).value).strip()  # 第3列是姓名列
                if cell_value and cell_value != "姓名":  # 排除表头和空值
                    student_rows[cell_value] = row_idx
                    
        return student_rows
    
    def write_grade_to_excel(self, excel_path: str, student_name: str, homework_dir_name: str, grade: str) -> None:
        """
        将学生成绩写入Excel文件（根据作业目录名定位列号）

        Args:
            excel_path: Excel文件路径
            student_name: 学生姓名
            homework_dir_name: 作业目录名（如“第1次作业”）
            grade: 成绩（A/B/C/D/E）
        """
        logger.info("="*50)
        logger.info(f"开始写入成绩到Excel文件")
        logger.info(f"Excel文件路径: {excel_path}")
        logger.info(f"学生姓名: {student_name}")
        logger.info(f"作业目录名: {homework_dir_name}")
        logger.info(f"成绩: {grade}")
        logger.info("-"*30)
        try:
            if not os.path.exists(excel_path):
                logger.error(f"Excel文件不存在: {excel_path}")
                raise FileNotFoundError(f"找不到Excel文件: {excel_path}")
            if not os.access(excel_path, os.W_OK):
                logger.error(f"无权限写入Excel文件: {excel_path}")
                raise PermissionError(f"无权限写入Excel文件: {excel_path}")
            with open(excel_path, 'rb') as f:
                wb = openpyxl.load_workbook(f)
                ws = wb.active
                logger.info(f"成功打开Excel文件")
                logger.info(f"当前工作表: {ws.title}")
                logger.info(f"工作表维度: 行数={ws.max_row}, 列数={ws.max_column}")
                # 加载学生名单
                if excel_path not in self._student_cache:
                    logger.info("加载学生名单到缓存")
                    self._student_cache[excel_path] = self._load_student_list(Path(excel_path))
                student_rows = self._student_cache[excel_path]
                if student_name not in student_rows:
                    logger.warning(f"学生 {student_name} 在 {excel_path} 中未找到")
                    logger.info("当前学生名单:")
                    for name in student_rows.keys():
                        logger.info(f"- {name}")
                    return
                student_row = student_rows[student_name]
                logger.info(f"找到学生 {student_name} 在第 {student_row} 行")
                # 解析作业序号
                match = re.search(r'第([一二三四五六七八九十\d]+)次作业|作业([一二三四五六七八九十\d]+)', homework_dir_name)
                if not match:
                    logger.error(f"无法从目录名中提取作业序号: {homework_dir_name}")
                    raise ValueError(f"无法从目录名中提取作业序号: {homework_dir_name}")
                number = match.group(1) or match.group(2)
                if re.match(r'[一二三四五六七八九十]+', number):
                    chinese_numbers = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5,
                                      '六': 6, '七': 7, '八': 8, '九': 9, '十': 10}
                    homework_number = chinese_numbers.get(number, 0)
                else:
                    homework_number = int(number)
                col_idx = 3 + homework_number  # 第4列起
                logger.info(f"成绩将写入第 {col_idx} 列 (作业序号: {homework_number})")
                ws.cell(row=student_row, column=col_idx, value=grade)
                try:
                    wb.save(excel_path)
                    logger.info(f"成功保存Excel文件: {excel_path}")
                except Exception as e:
                    logger.error(f"保存Excel文件失败: {str(e)}")
                    raise
        except Exception as e:
            logger.error(f"写入成绩到Excel文件失败: {str(e)}")
            raise
        finally:
            logger.info("="*50)
    
    def write_grade_to_excel_by_dirname(self, excel_path: str, student_name: str, homework_dir_name: str, grade: str) -> None:
        """
        新增：根据作业目录名直接定位成绩列（第4列起），不考虑列名。
        Args:
            excel_path: Excel文件路径
            student_name: 学生姓名
            homework_dir_name: 作业目录名（如“第1次作业”）
            grade: 成绩（A/B/C/D/E）
        """
        logger.info("="*50)
        logger.info(f"[ByDirName] 开始写入成绩到Excel文件")
        logger.info(f"Excel文件路径: {excel_path}")
        logger.info(f"学生姓名: {student_name}")
        logger.info(f"作业目录名: {homework_dir_name}")
        logger.info(f"成绩: {grade}")
        logger.info("-"*30)
        try:
            if not os.path.exists(excel_path):
                logger.error(f"Excel文件不存在: {excel_path}")
                raise FileNotFoundError(f"找不到Excel文件: {excel_path}")
            if not os.access(excel_path, os.W_OK):
                logger.error(f"无权限写入Excel文件: {excel_path}")
                raise PermissionError(f"无权限写入Excel文件: {excel_path}")
            with open(excel_path, 'rb') as f:
                wb = openpyxl.load_workbook(f)
                ws = wb.active
                logger.info(f"成功打开Excel文件")
                logger.info(f"当前工作表: {ws.title}")
                logger.info(f"工作表维度: 行数={ws.max_row}, 列数={ws.max_column}")
                if excel_path not in self._student_cache:
                    logger.info("加载学生名单到缓存")
                    self._student_cache[excel_path] = self._load_student_list(Path(excel_path))
                student_rows = self._student_cache[excel_path]
                if student_name not in student_rows:
                    logger.warning(f"学生 {student_name} 在 {excel_path} 中未找到")
                    logger.info("当前学生名单:")
                    for name in student_rows.keys():
                        logger.info(f"- {name}")
                    return
                student_row = student_rows[student_name]
                logger.info(f"找到学生 {student_name} 在第 {student_row} 行")
                match = re.search(r'第([一二三四五六七八九十\d]+)次作业|作业([一二三四五六七八九十\d]+)', homework_dir_name)
                if not match:
                    logger.error(f"无法从目录名中提取作业序号: {homework_dir_name}")
                    raise ValueError(f"无法从目录名中提取作业序号: {homework_dir_name}")
                number = match.group(1) or match.group(2)
                if re.match(r'[一二三四五六七八九十]+', number):
                    chinese_numbers = {'一': 1, '二': 2, '三': 3, '四': 4, '五': 5,
                                      '六': 6, '七': 7, '八': 8, '九': 9, '十': 10}
                    homework_number = chinese_numbers.get(number, 0)
                else:
                    homework_number = int(number)
                col_idx = 3 + homework_number  # 第4列起
                logger.info(f"成绩将写入第 {col_idx} 列 (作业序号: {homework_number})")
                ws.cell(row=student_row, column=col_idx, value=grade)
                try:
                    wb.save(excel_path)
                    logger.info(f"成功保存Excel文件: {excel_path}")
                except Exception as e:
                    logger.error(f"保存Excel文件失败: {str(e)}")
                    raise
        except Exception as e:
            logger.error(f"写入成绩到Excel文件失败: {str(e)}")
            raise
        finally:
            logger.info("="*50)
    
    def _extract_class_name(self, path: Path, is_multi_class: bool) -> str:
        """从文件路径中提取班级名称
        
        Args:
            path: 文件路径
            is_multi_class: 是否为多班级仓库
            
        Returns:
            str: 班级名称
            
        Raises:
            ValueError: 如果无法从路径中提取班级名称
        """
        # 如果是多班级仓库，从目录名中提取班级信息
        if is_multi_class:
            # 获取文件所在目录
            current_dir = path.parent
            
            # 检查目录名是否符合班级命名规范（例如：23计算机1班）
            class_pattern = re.compile(r'^\d{2}计算机[12]班$')
            if class_pattern.match(current_dir.name):
                return current_dir.name
                
            # 如果目录名不符合规范，尝试从路径中提取
            class_match = re.search(r'(\d{2}计算机[12]班)', str(path))
            if class_match:
                return class_match.group(1)
                
            # 如果还是找不到，尝试从父目录中查找
            for parent in current_dir.parents:
                if class_pattern.match(parent.name):
                    return parent.name
                    
            raise ValueError(f"Could not determine class name from path: {path}")
            
        # 如果是单班级仓库，使用默认班级名称
        return "22计算机G1班"
    
    def _extract_student_name(self, file_path: str) -> str:
        """从文件路径中提取学生姓名"""
        file_name = os.path.basename(file_path)
        # 移除文件扩展名
        name = os.path.splitext(file_name)[0]
        
        return name
    
    def _extract_grade_from_docx(self, path: Path) -> str:
        """
        Extract grade from docx file content.
        
        Args:
            path: Path to the docx file
            
        Returns:
            Grade as string (A, B, C, D, or E)
            
        Raises:
            ValueError: If an invalid grade is found
        """
        try:
            doc = Document(str(path))
            found_grade_pattern = False
            
            for paragraph in doc.paragraphs:
                if "老师评分：" in paragraph.text:
                    found_grade_pattern = True
                    grade_text = paragraph.text.split("老师评分：")[-1].strip()
                    if grade_text:
                        if grade_text in ['A', 'B', 'C', 'D', 'E']:
                            return grade_text
                        else:
                            raise ValueError(f"Invalid grade '{grade_text}' in file: {path}")
            
            # 如果没有找到成绩，返回默认成绩B
            return "B"
            
        except Exception as e:
            # 如果处理文件时出错，返回默认成绩B
            print(f"Error reading file {path}: {str(e)}")
            return "B"
    
    def process_docx_files(self, repository_path: str) -> None:
        """
        处理仓库中的所有docx文件，提取成绩并写入Excel

        Args:
            repository_path: 仓库路径
        """
        logger.info("="*50)
        logger.info(f"开始处理仓库: {repository_path}")
        logger.info("-"*30)
        
        repo_path = Path(repository_path)
        
        # 搜索成绩登记表文件
        excel_files = list(repo_path.glob("平时成绩登记表-*.xlsx"))
        if not excel_files:
            logger.error(f"在 {repository_path} 中未找到成绩登记表文件")
            raise FileNotFoundError(f"在 {repository_path} 中未找到成绩登记表文件")

        logger.info(f"找到 {len(excel_files)} 个成绩登记表文件:")
        for excel_file in excel_files:
            logger.info(f"- {excel_file}")
        
        # 处理每个成绩登记表文件
        for excel_file in excel_files:
            logger.info("-"*30)
            logger.info(f"开始处理成绩登记表: {excel_file}")

            # 从文件名中提取专业和年级信息
            index = excel_file.stem.find("-")
            class_name = excel_file.stem[index + 1:]# 例如：23计算机1-2班 或 22计算机G1班
            # class_name = excel_file.stem.split("-")[-1]  
            logger.info(f"class_name: {class_name}")
            classes = class_name.split("班")[0].split("计算机")[1]  # 例如：1-2 或 G1
            year = class_name.split("计算机")[0]  # 例如：23
            
            logger.info(f"班级信息:")
            logger.info(f"- 年级: {year}")
            logger.info(f"- 班级: {classes}")

            # 根据成绩登记表文件名判断是单班级还是多班级
            is_multi_class = "-" in classes
            logger.info(f"根据成绩登记表判断: {'多班级' if is_multi_class else '单班级'}")

            if is_multi_class:
                self._process_multi_class(str(repo_path))
            else:
                self._process_single_class(str(repo_path), str(excel_file), year, classes)
            
        logger.info("="*50)

    def _process_single_class(self, repo_path: str, excel_file: str, year: str, major: str) -> None:
        """处理单班级仓库"""
        logger.info("处理单班级仓库")
        logger.info("=" * 50)
        
        # 查找所有作业目录
        homework_dirs = []
        for item in os.listdir(repo_path):
            item_path = os.path.join(repo_path, item)
            if os.path.isdir(item_path) and re.match(r'第[一二三四五六七八九十]+次作业|作业[一二三四五六七八九十]+|作业\d+', item):
                homework_dirs.append(item_path)
        
        if not homework_dirs:
            logger.warning(f"在仓库中未找到作业目录: {repo_path}")
            return
        
        logger.info(f"找到 {len(homework_dirs)} 个作业目录")
        
        # 处理每个作业目录
        for homework_dir in homework_dirs:
            logger.info(f"处理作业目录: {homework_dir}")
            self._process_docx_files_batch(homework_dir, excel_file)

    def _is_multi_class_repo(self, repo_path: str) -> bool:
        """判断仓库是否为多班级仓库
        
        Args:
            repo_path: 仓库根目录路径
            
        Returns:
            bool: 如果是多班级仓库返回True，否则返回False
        """
        # 查找成绩登记表
        excel_files = glob.glob(os.path.join(repo_path, "平时成绩登记表-*.xlsx"))
        if not excel_files:
            return False
        
        # 检查文件名是否包含多班级信息（如：23计算机1-2班）
        for excel_file in excel_files:
            file_name = os.path.basename(excel_file)
            if re.search(r'计算机\d+-\d+班', file_name):
                return True
            
        return False

    def _process_multi_class(self, repo_path: str) -> None:
        """处理多班级仓库"""
        logger.info("处理多班级仓库")
        logger.info("=" * 50)
        
        # 查找所有成绩登记表
        excel_files = glob.glob(os.path.join(repo_path, "平时成绩登记表-*.xlsx"))
        if not excel_files:
            raise ValueError(f"在仓库 {repo_path} 中未找到成绩登记表")
        
        logger.info(f"找到 {len(excel_files)} 个成绩登记表文件:")
        for excel_file in excel_files:
            logger.info(f"- {excel_file}")
        
        # 处理每个成绩登记表
        for excel_file in excel_files:
            logger.info("-" * 30)
            logger.info(f"开始处理成绩登记表: {excel_file}")
            
            # 从文件名中提取班级信息
            file_name = os.path.basename(excel_file)
            match = re.match(r'平时成绩登记表-(\d{2})计算机(\d+)-(\d+)班\.xlsx', file_name)
            if not match:
                logger.warning(f"无法从文件名中提取班级信息: {file_name}")
                continue
            
            year, start_class, end_class = match.groups()
            logger.info("班级信息:")
            logger.info(f"- 年级: {year}")
            logger.info(f"- 班级范围: {start_class}-{end_class}")
            
            # 查找对应的班级目录
            class_dirs = []
            for class_num in range(int(start_class), int(end_class) + 1):
                class_dir = os.path.join(repo_path, f"{year}计算机{class_num}班")
                if os.path.exists(class_dir):
                    class_dirs.append(class_dir)
            
            if not class_dirs:
                logger.warning(f"未找到对应的班级目录")
                continue
            
            # 处理每个班级目录下的作业
            for class_dir in class_dirs:
                self._process_single_class(class_dir, excel_file, year, "")
                # logger.info(f"处理班级目录: {class_dir}")
                # self._process_docx_files_batch(class_dir, excel_file)

    def _process_docx_files_batch(self, dir_path: str, excel_file: str) -> None:
        """处理指定目录下的所有docx文件"""
        docx_files = glob.glob(os.path.join(dir_path, "*.docx"))
        if not docx_files:
            logger.warning(f"在 {dir_path} 中未找到docx文件")
            return
        logger.info(f"在 {dir_path} 中找到 {len(docx_files)} 个docx文件")
        homework_dir_name = os.path.basename(dir_path)
        for docx_file in docx_files:
            try:
                if not self._is_valid_docx(docx_file):
                    logger.warning(f"文件格式无效: {docx_file}")
                    continue
                logger.info(f"处理文件: {docx_file}")
                logger.info("=" * 50)
                try:
                    student_name = self._extract_student_name(docx_file)
                except ValueError as e:
                    logger.warning(str(e))
                    continue
                try:
                    grade = self._extract_grade_from_docx(Path(docx_file))
                except Exception as e:
                    logger.warning(f"提取成绩失败: {str(e)}")
                    continue
                logger.info("提取信息:")
                logger.info(f"- 学生: {student_name}")
                logger.info(f"- 作业目录: {homework_dir_name}")
                logger.info(f"- 成绩: {grade}")
                self.write_grade_to_excel(excel_file, student_name, homework_dir_name, grade)
            except Exception as e:
                logger.error(f"处理文件 {docx_file} 时发生错误: {str(e)}")
                continue

    def _is_valid_docx(self, file_path: str) -> bool:
        """验证文件是否为有效的docx文件"""
        try:
            with zipfile.ZipFile(file_path, 'r') as zip_ref:
                # 检查是否包含必要的docx文件结构
                required_files = ['word/document.xml', '[Content_Types].xml']
                return all(file in zip_ref.namelist() for file in required_files)
        except:
            return False 