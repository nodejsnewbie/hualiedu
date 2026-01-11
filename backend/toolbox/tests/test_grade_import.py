from pathlib import Path
from tempfile import TemporaryDirectory

from django.test import SimpleTestCase
from openpyxl import Workbook, load_workbook

from toolbox.utils import (
    AssignmentImportError,
    import_assignment_scores_to_gradebook,
)


class AssignmentGradeImportTests(SimpleTestCase):
    def setUp(self):
        super().setUp()
        self.tmp_dir = TemporaryDirectory()
        self.addCleanup(self.tmp_dir.cleanup)
        self.tmp_path = Path(self.tmp_dir.name)

    def test_scores_written_to_expected_column(self):
        assignment_file = self._create_assignment_workbook()
        gradebook_file = self._create_gradebook_workbook()

        result = import_assignment_scores_to_gradebook(
            assignment_file, gradebook_file, assignment_number=5
        )

        self.assertEqual(result["updated_students"], 2)
        self.assertEqual(result["assignment_column_letter"], "H")
        self.assertEqual(result["missing_in_gradebook"], [])
        self.assertEqual(len(result["missing_in_assignment"]), 0)

        workbook = load_workbook(gradebook_file)
        worksheet = workbook.active
        self.assertEqual(worksheet.cell(row=4, column=8).value, 95.0)
        self.assertEqual(worksheet.cell(row=5, column=8).value, 88.0)
        workbook.close()

    def test_invalid_assignment_number_raises(self):
        assignment_file = self._create_assignment_workbook()
        gradebook_file = self._create_gradebook_workbook()

        with self.assertRaises(AssignmentImportError):
            import_assignment_scores_to_gradebook(
                assignment_file, gradebook_file, assignment_number=0
            )

    def _create_assignment_workbook(self) -> str:
        file_path = self.tmp_path / "assignment.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["作业名称：第五次作业"])
        worksheet.append(["学号/工号", "学生姓名", "单位ID", "课程名称", "分数"])
        worksheet.append([None, None, None, None, None])
        worksheet.append(["1001", "张三", "1", "课程A", 95])
        worksheet.append(["1002", "李四", "1", "课程A", 88])
        workbook.save(file_path)
        workbook.close()
        return str(file_path)

    def _create_gradebook_workbook(self) -> str:
        file_path = self.tmp_path / "gradebook.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.append(["广州华立学院学生平时成绩登记表"])
        worksheet.append(["序号", "学号", "姓名", "平时成绩", None, None, None, None])
        worksheet.append([None, None, None, None, None, None, None, None])
        worksheet.append(["1", "1001", "张三", None, None, None, None, None])
        worksheet.append(["2", "1002", "李四", None, None, None, None, None])
        workbook.save(file_path)
        workbook.close()
        return str(file_path)
