import os
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class AssignmentImportError(Exception):
    """Raised when the grade import process cannot be completed."""


def import_assignment_scores_to_gradebook(
    assignment_file: str,
    gradebook_file: str,
    assignment_number: int,
    sheet_name: str = "Sheet1",
) -> Dict[str, object]:
    """
    Copy scores from an assignment workbook into the gradebook workbook.

    Args:
        assignment_file: Path to the assignment workbook (Excel).
        gradebook_file: Path to the gradebook workbook (Excel).
        assignment_number: The assignment index (1-based). The first assignment
            is written to the first score column after the "姓名" column, the
            second assignment to the next column, and so on.
        sheet_name: Target sheet inside the gradebook workbook.

    Returns:
        A dict containing statistics about the import process.
    """

    if assignment_number < 1:
        raise AssignmentImportError("作业编号必须大于等于 1")

    assignment_path = _validate_excel_path(assignment_file, "作业成绩表")
    gradebook_path = _validate_excel_path(gradebook_file, "成绩登分册")

    assignment_wb = load_workbook(assignment_path, data_only=True)
    gradebook_wb = load_workbook(gradebook_path)

    try:
        assignment_ws = assignment_wb.active
        gradebook_ws = _get_gradebook_sheet(gradebook_wb, sheet_name)

        assignment_header_row, assignment_header_map = _locate_header_row(
            assignment_ws, {"学生姓名", "分数"}
        )
        gradebook_header_row, gradebook_header_map = _locate_header_row(
            gradebook_ws, {"姓名"}
        )

        assignment_scores = _collect_assignment_scores(
            assignment_ws, assignment_header_row, assignment_header_map
        )
        if not assignment_scores:
            raise AssignmentImportError("未在作业成绩表中找到任何有效成绩")

        gradebook_rows = _collect_gradebook_rows(
            gradebook_ws, gradebook_header_row, gradebook_header_map["姓名"]
        )
        if not gradebook_rows:
            raise AssignmentImportError("成绩登分册中没有可写入的学生信息")

        target_column_index = gradebook_header_map["姓名"] + assignment_number
        target_column_letter = get_column_letter(target_column_index)

        updates = _write_scores_to_gradebook(
            gradebook_ws, gradebook_rows, assignment_scores, target_column_index
        )

        missing_in_gradebook = sorted(
            set(assignment_scores.keys()) - set(gradebook_rows.keys())
        )
        missing_in_assignment = sorted(
            set(gradebook_rows.keys()) - set(assignment_scores.keys())
        )

        gradebook_wb.save(gradebook_path)

        return {
            "assignment_file": str(assignment_path),
            "gradebook_file": str(gradebook_path),
            "assignment_number": assignment_number,
            "assignment_column_letter": target_column_letter,
            "assignment_column_index": target_column_index,
            "updated_students": len(updates),
            "updates": updates,
            "missing_in_gradebook": missing_in_gradebook,
            "missing_in_assignment": missing_in_assignment,
        }
    finally:
        assignment_wb.close()
        gradebook_wb.close()


def _validate_excel_path(path_str: str, label: str) -> Path:
    path = Path(os.path.expanduser(path_str)).resolve()
    if not path.exists():
        raise AssignmentImportError(f"{label}文件不存在：{path}")
    if path.suffix.lower() not in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        raise AssignmentImportError(f"{label}文件格式不受支持：{path.suffix}")
    return path


def _get_gradebook_sheet(workbook, sheet_name: str) -> Worksheet:
    try:
        return workbook[sheet_name]
    except KeyError as exc:
        raise AssignmentImportError(f"成绩登分册中找不到工作表：{sheet_name}") from exc


def _locate_header_row(
    worksheet: Worksheet, required_headers: Iterable[str], max_search_rows: int = 20
) -> Tuple[int, Dict[str, int]]:
    required = {header.strip() for header in required_headers}
    for row in worksheet.iter_rows(min_row=1, max_row=max_search_rows):
        header_map: Dict[str, int] = {}
        for idx, cell in enumerate(row, start=1):
            value = _normalize_cell_value(cell.value)
            if value:
                header_map[value] = idx
        if required.issubset(header_map.keys()):
            return row[0].row, header_map

    raise AssignmentImportError(f"无法在工作表中找到表头：{', '.join(required)}")


def _collect_assignment_scores(
    worksheet: Worksheet, header_row: int, header_map: Dict[str, int]
) -> Dict[str, float]:
    name_col = header_map["学生姓名"]
    score_col = header_map["分数"]
    scores: Dict[str, float] = {}

    for row in worksheet.iter_rows(min_row=header_row + 1):
        name_value = _normalize_cell_value(row[name_col - 1].value)
        if not name_value:
            continue
        score_cell = row[score_col - 1].value
        if score_cell is None or score_cell == "":
            continue
        try:
            score = float(score_cell)
        except (TypeError, ValueError):
            raise AssignmentImportError(f"无法解析学生 {name_value} 的成绩：{score_cell}")
        scores[name_value] = score

    return scores


def _collect_gradebook_rows(
    worksheet: Worksheet, header_row: int, name_column: int
) -> Dict[str, int]:
    rows: Dict[str, int] = {}
    for row in worksheet.iter_rows(min_row=header_row + 1):
        name_value = _normalize_cell_value(row[name_column - 1].value)
        if not name_value or name_value == "例：":
            continue
        rows[name_value] = row[name_column - 1].row
    return rows


def _write_scores_to_gradebook(
    worksheet: Worksheet,
    gradebook_rows: Dict[str, int],
    assignment_scores: Dict[str, float],
    target_column: int,
) -> List[Dict[str, object]]:
    updates: List[Dict[str, object]] = []
    for name, row_index in gradebook_rows.items():
        if name not in assignment_scores:
            continue
        score = assignment_scores[name]
        worksheet.cell(row=row_index, column=target_column, value=score)
        updates.append({"name": name, "row": row_index, "score": score})
    return updates


def _normalize_cell_value(value: Optional[object]) -> Optional[str]:
    if value is None:
        return None
    if isinstance(value, str):
        return value.strip() or None
    return str(value).strip() or None
