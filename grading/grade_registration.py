import glob
import logging
import os
import re
from pathlib import Path
from typing import Dict

import openpyxl
from docx import Document

logger = logging.getLogger(__name__)


class GradeRegistration:
    def __init__(self) -> None:
        self.repo_path: Path = Path(".")
        self._student_cache: Dict[str, Dict[str, int]] = {}

    # --------------------- helpers ---------------------
    def _ensure_excel_headers(self, excel_path: str) -> openpyxl.workbook.workbook.Workbook:
        if not os.path.exists(excel_path):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.cell(row=1, column=1, value="序号")
            ws.cell(row=1, column=2, value="学号")
            ws.cell(row=1, column=3, value="姓名")
            wb.save(excel_path)
        with open(excel_path, "rb") as f:
            return openpyxl.load_workbook(f)

    def _load_student_list(self, excel_path: Path) -> Dict[str, int]:
        student_rows: Dict[str, int] = {}
        if not excel_path.exists():
            return student_rows
        with open(str(excel_path), "rb") as f:
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            for row_idx in range(1, ws.max_row + 1):
                value = ws.cell(row=row_idx, column=3).value
                if value is None:
                    continue
                name = str(value).strip()
                if name and name != "姓名":
                    student_rows[name] = row_idx
        return student_rows

    def _parse_homework_number(self, homework_dir_name: str) -> int:
        m = re.search(
            r"第([一二三四五六七八九十\d]+)次作业|作业([一二三四五六七八九十\d]+)",
            homework_dir_name,
        )
        number = (m.group(1) or m.group(2)) if m else "1"
        if re.fullmatch(r"[一二三四五六七八九十]+", number):
            mapping = {
                "一": 1,
                "二": 2,
                "三": 3,
                "四": 4,
                "五": 5,
                "六": 6,
                "七": 7,
                "八": 8,
                "九": 9,
                "十": 10,
            }
            return mapping.get(number, 1)
        return int(number)

    def _extract_student_name(self, file_path: str) -> str:
        file_name = os.path.basename(file_path)
        name = os.path.splitext(file_name)[0]
        # 去除常见分隔符后的说明性后缀
        name = re.split(r"[_\-\s]", name)[0]
        # 移除数字
        name = re.sub(r"\d+", "", name)
        # 仅保留中文字符（如果为空，退回原始）
        chinese_only = "".join(ch for ch in name if "\u4e00" <= ch <= "\u9fff")
        return chinese_only or name

    def _extract_grade_from_docx(self, path: Path) -> str:
        try:
            doc = Document(str(path))
            for p in doc.paragraphs:
                if "老师评分：" in p.text:
                    g = p.text.split("老师评分：")[-1].strip()
                    if g in ["A", "B", "C", "D", "E"]:
                        return g
                    raise ValueError(f"Invalid grade '{g}' in file: {path}")
            return "B"
        except Exception:
            return "B"

    # --------------------- public APIs ---------------------
    def write_grade_to_excel(
        self, excel_path: str, student_name: str, homework_dir_name: str, grade: str
    ) -> None:
        wb = self._ensure_excel_headers(excel_path)
        ws = wb.active

        if excel_path not in self._student_cache:
            self._student_cache[excel_path] = self._load_student_list(Path(excel_path))
        student_rows = self._student_cache[excel_path]

        if student_name not in student_rows:
            new_row = ws.max_row + 1
            ws.cell(row=new_row, column=1, value=new_row - 1)
            ws.cell(row=new_row, column=3, value=student_name)
            student_rows[student_name] = new_row
        row_idx = student_rows[student_name]

        hw_no = self._parse_homework_number(homework_dir_name)
        col_idx = 3 + hw_no
        if ws.cell(row=1, column=col_idx).value in (None, ""):
            # 规范化表头为“第N次作业”
            ws.cell(row=1, column=col_idx, value=f"第{hw_no}次作业")
        ws.cell(row=row_idx, column=col_idx, value=grade)

        wb.save(excel_path)

    def _is_multi_class_repo(self, repo_path: str) -> bool:
        excel_files = glob.glob(os.path.join(repo_path, "平时成绩登记表-*.xlsx"))
        for p in excel_files:
            if re.search(r"平时成绩登记表-\d{2}计算机\d+-\d+班\.xlsx$", os.path.basename(p)):
                return True
        return False

    def process_docx_files(self, repository_path: str) -> None:
        repo_path = Path(repository_path)
        if not repo_path.exists():
            logger.warning(f"仓库路径不存在: {repository_path}")
            return

        # 查找根目录下的登记表
        excel_files = list(repo_path.glob("平时成绩登记表-*.xlsx"))
        if not excel_files:
            logger.warning(f"在 {repository_path} 中未找到成绩登记表文件")
            return
        excel_path = str(excel_files[0])

        # 遍历docx
        for docx_file in repo_path.rglob("*.docx"):
            # 最近包含“作业”的父目录名
            hw_dir_name = docx_file.parent.name
            for parent in docx_file.parents:
                if "作业" in parent.name:
                    hw_dir_name = parent.name
                    break
            try:
                student_name = self._extract_student_name(str(docx_file))
                grade = self._extract_grade_from_docx(docx_file)
                self.write_grade_to_excel(excel_path, student_name, hw_dir_name, grade)
            except Exception as e:
                logger.warning(f"处理文件失败 {docx_file}: {e}")
