"""
作业成绩写入成绩登分册服务层

提供两种场景的成绩写入服务：
1. 作业评分系统场景：从作业目录批量写入成绩
2. 工具箱模块场景：从班级目录的Excel文件批量写入成绩
"""

import os
import logging
import math
from datetime import datetime
from typing import Dict, List, Optional, Tuple

from django.core.cache import cache
from django.utils import timezone
from grading.grade_registry_writer import (
    GradeFileProcessor,
    RegistryManager,
    NameMatcher,
)

logger = logging.getLogger(__name__)


class BatchGradeProgressTracker:
    """批量登分进度跟踪器"""

    CACHE_KEY_PREFIX = "batch_grade_progress"
    CACHE_TIMEOUT_SECONDS = 60 * 30

    def __init__(self, tracking_id: str, user_id: Optional[int] = None):
        self.tracking_id = tracking_id
        self.user_id = user_id
        self.state: Dict[str, any] = {
            "tracking_id": tracking_id,
            "user_id": user_id,
            "status": "pending",
            "total": 0,
            "processed": 0,
            "success": 0,
            "failed": 0,
            "skipped": 0,
            "current_file": None,
            "message": "等待开始",
            "started_at": timezone.now().isoformat(),
            "updated_at": timezone.now().isoformat(),
        }

    @classmethod
    def cache_key(cls, tracking_id: str) -> str:
        return f"{cls.CACHE_KEY_PREFIX}:{tracking_id}"

    def _save(self):
        self.state["updated_at"] = timezone.now().isoformat()
        cache.set(
            self.cache_key(self.tracking_id),
            self.state,
            self.CACHE_TIMEOUT_SECONDS,
        )

    def start(self, total_files: int = 0, message: str = "准备中..."):
        self.state.update(
            {
                "status": "preparing",
                "total": max(total_files, 0),
                "processed": 0,
                "success": 0,
                "failed": 0,
                "skipped": 0,
                "current_file": None,
                "message": message,
            }
        )
        self._save()

    def update_total(self, total_files: int):
        self.state["total"] = max(total_files, 0)
        self._save()

    def update_progress(
        self,
        *,
        processed: int,
        success: int,
        failed: int,
        skipped: int,
        current_file: Optional[str] = None,
        message: Optional[str] = None,
    ):
        self.state.update(
            {
                "status": "running",
                "processed": max(processed, 0),
                "success": max(success, 0),
                "failed": max(failed, 0),
                "skipped": max(skipped, 0),
                "current_file": current_file,
            }
        )
        if message:
            self.state["message"] = message
        elif current_file:
            self.state["message"] = f"正在处理 {current_file}"
        else:
            self.state["message"] = "正在批量登分..."
        self._save()

    def complete(self, summary: Optional[Dict[str, int]] = None, message: str = "批量登分完成"):
        self.state.update(
            {
                "status": "success",
                "message": message,
                "completed_at": timezone.now().isoformat(),
            }
        )
        if summary:
            self.state["final_summary"] = summary
            self.state["processed"] = summary.get("total", self.state["processed"])
            self.state["success"] = summary.get("success", self.state["success"])
            self.state["failed"] = summary.get("failed", self.state["failed"])
            self.state["skipped"] = summary.get("skipped", self.state["skipped"])
        self._save()

    def fail(self, message: str):
        self.state.update(
            {
                "status": "error",
                "message": message,
                "completed_at": timezone.now().isoformat(),
            }
        )
        self._save()

    @classmethod
    def get_progress(cls, tracking_id: str) -> Optional[Dict[str, any]]:
        if not tracking_id:
            return None
        return cache.get(cls.cache_key(tracking_id))


class AuditLogger:
    """操作审计日志记录器"""

    def __init__(self, user, tenant, scenario: str):
        """
        初始化审计日志记录器

        Args:
            user: 当前用户
            tenant: 当前租户
            scenario: 使用场景
        """
        self.user = user
        self.tenant = tenant
        self.scenario = scenario
        self.logger = logging.getLogger(f"{__name__}.audit")
        self.operation_start_time = None
        self.audit_data = {
            "user_id": user.id if user else None,
            "username": user.username if user else None,
            "tenant_id": tenant.id if tenant else None,
            "tenant_name": tenant.name if tenant else None,
            "scenario": scenario,
            "start_time": None,
            "end_time": None,
            "duration_seconds": None,
        }

    def start_operation(self, operation_type: str, **kwargs):
        """
        记录操作开始

        Args:
            operation_type: 操作类型
            **kwargs: 额外的审计信息
        """
        self.operation_start_time = datetime.now()
        self.audit_data["start_time"] = self.operation_start_time.isoformat()
        self.audit_data["operation_type"] = operation_type
        self.audit_data.update(kwargs)

        self.logger.info(
            "操作开始 - 用户: %s (ID: %s), 租户: %s (ID: %s), 场景: %s, 操作: %s",
            self.audit_data["username"],
            self.audit_data["user_id"],
            self.audit_data["tenant_name"],
            self.audit_data["tenant_id"],
            self.scenario,
            operation_type,
        )

        # 记录额外的审计信息
        for key, value in kwargs.items():
            self.logger.info("  %s: %s", key, value)

    def end_operation(self, success: bool, **kwargs):
        """
        记录操作结束

        Args:
            success: 操作是否成功
            **kwargs: 额外的审计信息
        """
        end_time = datetime.now()
        self.audit_data["end_time"] = end_time.isoformat()
        self.audit_data["success"] = success

        if self.operation_start_time:
            duration = (end_time - self.operation_start_time).total_seconds()
            self.audit_data["duration_seconds"] = duration
        else:
            duration = None

        self.audit_data.update(kwargs)

        status = "成功" if success else "失败"
        self.logger.info(
            "操作结束 - 用户: %s, 租户: %s, 场景: %s, 状态: %s, 耗时: %.2f秒",
            self.audit_data["username"],
            self.audit_data["tenant_name"],
            self.scenario,
            status,
            duration if duration else 0,
        )

        # 记录统计信息
        for key, value in kwargs.items():
            self.logger.info("  %s: %s", key, value)

    def log_grade_write(
        self, student_name: str, homework_number: int, grade: str, old_grade: Optional[str] = None
    ):
        """
        记录成绩写入操作

        Args:
            student_name: 学生姓名
            homework_number: 作业批次
            grade: 新成绩
            old_grade: 旧成绩（如果有）
        """
        if old_grade and old_grade != grade:
            self.logger.info(
                "成绩覆盖 - 学生: %s, 作业: 第%d次, 旧成绩: %s, 新成绩: %s",
                student_name,
                homework_number,
                old_grade,
                grade,
            )
        else:
            self.logger.info(
                "成绩写入 - 学生: %s, 作业: 第%d次, 成绩: %s",
                student_name,
                homework_number,
                grade,
            )

    def log_file_processing(self, file_path: str, status: str, error: Optional[str] = None):
        """
        记录文件处理状态

        Args:
            file_path: 文件路径
            status: 处理状态 (success/failed/skipped)
            error: 错误信息（如果有）
        """
        if status == "success":
            self.logger.info("文件处理成功: %s", file_path)
        elif status == "failed":
            self.logger.warning("文件处理失败: %s, 原因: %s", file_path, error)
        elif status == "skipped":
            self.logger.debug("文件跳过: %s, 原因: %s", file_path, error)


class GradeRegistryWriterService:
    """作业成绩写入服务"""

    # 场景常量
    SCENARIO_GRADING_SYSTEM = "grading_system"
    SCENARIO_TOOLBOX = "toolbox"
    
    # 性能优化：文件数量限制
    MAX_FILES_WARNING_THRESHOLD = 500
    
    # 安全加固：文件大小限制（100MB）
    MAX_FILE_SIZE = 100 * 1024 * 1024

    def __init__(self, user, tenant, scenario: str):
        """
        初始化服务

        Args:
            user: 当前用户
            tenant: 当前租户
            scenario: 使用场景 (SCENARIO_GRADING_SYSTEM 或 SCENARIO_TOOLBOX)
        """
        self.user = user
        self.tenant = tenant
        self.scenario = scenario
        self.logger = logging.getLogger(__name__)
        self.audit_logger = AuditLogger(user, tenant, scenario)

        # 验证场景
        if scenario not in [self.SCENARIO_GRADING_SYSTEM, self.SCENARIO_TOOLBOX]:
            raise ValueError(f"无效的场景类型: {scenario}")

        self.logger.info(
            "初始化GradeRegistryWriterService - 用户: %s, 租户: %s, 场景: %s",
            user.username if user else "None",
            tenant.name if tenant else "None",
            scenario,
        )

    def _validate_path_security(self, file_path: str, base_dir: str) -> Tuple[bool, Optional[str]]:
        """
        强化路径验证逻辑（安全加固）

        Args:
            file_path: 文件路径
            base_dir: 基础目录

        Returns:
            (是否有效, 错误消息)
        """
        try:
            # 1. 检查路径遍历攻击
            abs_file_path = os.path.abspath(file_path)
            abs_base_dir = os.path.abspath(base_dir)

            if not abs_file_path.startswith(abs_base_dir):
                self.logger.error(
                    "安全检查失败：路径遍历攻击检测 - 文件: %s, 基础目录: %s",
                    file_path,
                    base_dir
                )
                self.audit_logger.logger.warning(
                    "安全事件：路径遍历攻击尝试 - 用户: %s, 租户: %s, 文件: %s",
                    self.user.username if self.user else "Unknown",
                    self.tenant.name if self.tenant else "Unknown",
                    file_path
                )
                return False, "无权访问该路径"

            # 2. 检查路径中是否包含危险字符和模式
            dangerous_patterns = ["../", "..\\", "~", "$", "|", "&", ";", "`"]
            for pattern in dangerous_patterns:
                if pattern in file_path:
                    self.logger.error(
                        "安全检查失败：路径包含危险字符 - 文件: %s, 模式: %s",
                        file_path,
                        pattern
                    )
                    self.audit_logger.logger.warning(
                        "安全事件：危险路径字符检测 - 用户: %s, 租户: %s, 文件: %s, 模式: %s",
                        self.user.username if self.user else "Unknown",
                        self.tenant.name if self.tenant else "Unknown",
                        file_path,
                        pattern
                    )
                    return False, "路径包含非法字符"

            # 3. 检查文件是否存在
            if not os.path.exists(file_path):
                return False, f"文件不存在: {file_path}"

            # 4. 检查是否是符号链接（防止符号链接攻击）
            if os.path.islink(file_path):
                self.logger.error("安全检查失败：符号链接检测 - 文件: %s", file_path)
                self.audit_logger.logger.warning(
                    "安全事件：符号链接检测 - 用户: %s, 租户: %s, 文件: %s",
                    self.user.username if self.user else "Unknown",
                    self.tenant.name if self.tenant else "Unknown",
                    file_path
                )
                return False, "不允许访问符号链接"

            # 5. 检查文件权限
            if not os.access(file_path, os.R_OK):
                self.logger.error("安全检查失败：无读取权限 - 文件: %s", file_path)
                self.audit_logger.logger.warning(
                    "安全事件：无权限访问文件 - 用户: %s, 租户: %s, 文件: %s",
                    self.user.username if self.user else "Unknown",
                    self.tenant.name if self.tenant else "Unknown",
                    file_path
                )
                return False, "无权限读取文件"

            # 6. 检查文件类型（只允许目录或常规文件）
            if not (os.path.isfile(file_path) or os.path.isdir(file_path)):
                self.logger.error("安全检查失败：非法文件类型 - 文件: %s", file_path)
                self.audit_logger.logger.warning(
                    "安全事件：非法文件类型 - 用户: %s, 租户: %s, 文件: %s",
                    self.user.username if self.user else "Unknown",
                    self.tenant.name if self.tenant else "Unknown",
                    file_path
                )
                return False, "非法文件类型"

            return True, None

        except Exception as e:
            self.logger.error("路径安全验证失败: %s", str(e), exc_info=True)
            self.audit_logger.logger.error(
                "安全事件：路径验证异常 - 用户: %s, 租户: %s, 文件: %s, 错误: %s",
                self.user.username if self.user else "Unknown",
                self.tenant.name if self.tenant else "Unknown",
                file_path,
                str(e)
            )
            return False, f"路径验证失败: {str(e)}"

    def _validate_file_size(self, file_path: str) -> Tuple[bool, Optional[str]]:
        """
        验证文件大小（安全加固）

        Args:
            file_path: 文件路径

        Returns:
            (是否有效, 错误消息)
        """
        try:
            file_size = os.path.getsize(file_path)

            # 检查文件大小是否为0（可能是损坏的文件）
            if file_size == 0:
                self.logger.warning("文件大小为0 - 文件: %s", file_path)
                return False, "文件为空或损坏"

            # 检查文件大小是否超过限制
            if file_size > self.MAX_FILE_SIZE:
                self.logger.error(
                    "文件大小超过限制 - 文件: %s, 大小: %d bytes, 限制: %d bytes",
                    file_path,
                    file_size,
                    self.MAX_FILE_SIZE
                )
                self.audit_logger.logger.warning(
                    "安全事件：文件大小超限 - 用户: %s, 租户: %s, 文件: %s, 大小: %.2fMB",
                    self.user.username if self.user else "Unknown",
                    self.tenant.name if self.tenant else "Unknown",
                    file_path,
                    file_size / 1024 / 1024
                )
                return False, (
                    f"文件大小超过限制 "
                    f"({file_size / 1024 / 1024:.2f}MB > "
                    f"{self.MAX_FILE_SIZE / 1024 / 1024:.2f}MB)"
                )

            return True, None

        except Exception as e:
            self.logger.error("文件大小验证失败: %s", str(e), exc_info=True)
            return False, f"文件大小验证失败: {str(e)}"

    def _validate_excel_integrity(self, file_path: str) -> Tuple[bool, Optional[str]]:
        """
        验证Excel文件完整性（安全加固）

        Args:
            file_path: Excel文件路径

        Returns:
            (是否有效, 错误消息)
        """
        try:
            # 1. 检查文件扩展名
            if not file_path.lower().endswith(".xlsx"):
                self.logger.error("非Excel文件格式: %s", file_path)
                return False, "文件格式错误：必须是Excel文件(.xlsx)"

            # 2. 尝试打开Excel文件验证完整性
            from openpyxl import load_workbook

            workbook = load_workbook(file_path, read_only=True, data_only=True)
            worksheet = workbook.active

            # 3. 检查是否有有效的工作表
            if worksheet is None:
                self.logger.error("Excel文件没有有效的工作表: %s", file_path)
                workbook.close()
                return False, "Excel文件格式错误：没有有效的工作表"

            # 4. 检查是否有数据（至少要有表头行）
            if worksheet.max_row < 1:
                self.logger.warning("Excel文件没有数据: %s", file_path)
                workbook.close()
                return False, "Excel文件没有数据"

            # 5. 检查工作表是否有列
            if worksheet.max_column < 1:
                self.logger.warning("Excel文件没有列: %s", file_path)
                workbook.close()
                return False, "Excel文件格式错误：没有列"

            # 6. 检查是否有过多的行或列（防止恶意文件）
            MAX_ROWS = 10000
            MAX_COLS = 100
            if worksheet.max_row > MAX_ROWS:
                self.logger.error(
                    "Excel文件行数过多: %s, 行数: %d, 限制: %d",
                    file_path,
                    worksheet.max_row,
                    MAX_ROWS
                )
                workbook.close()
                self.audit_logger.logger.warning(
                    "安全事件：Excel文件行数异常 - 用户: %s, 租户: %s, 文件: %s, 行数: %d",
                    self.user.username if self.user else "Unknown",
                    self.tenant.name if self.tenant else "Unknown",
                    file_path,
                    worksheet.max_row
                )
                return False, f"Excel文件行数过多（{worksheet.max_row} > {MAX_ROWS}）"

            if worksheet.max_column > MAX_COLS:
                self.logger.error(
                    "Excel文件列数过多: %s, 列数: %d, 限制: %d",
                    file_path,
                    worksheet.max_column,
                    MAX_COLS
                )
                workbook.close()
                self.audit_logger.logger.warning(
                    "安全事件：Excel文件列数异常 - 用户: %s, 租户: %s, 文件: %s, 列数: %d",
                    self.user.username if self.user else "Unknown",
                    self.tenant.name if self.tenant else "Unknown",
                    file_path,
                    worksheet.max_column
                )
                return False, f"Excel文件列数过多（{worksheet.max_column} > {MAX_COLS}）"

            workbook.close()
            return True, None

        except PermissionError as e:
            self.logger.error("Excel文件权限错误: %s - %s", file_path, str(e))
            return False, "无权限访问Excel文件"
        except Exception as e:
            self.logger.error(
                "Excel文件完整性验证失败: %s - %s",
                file_path,
                str(e),
                exc_info=True
            )
            self.audit_logger.logger.warning(
                "安全事件：Excel文件验证失败 - 用户: %s, 租户: %s, 文件: %s, 错误: %s",
                self.user.username if self.user else "Unknown",
                self.tenant.name if self.tenant else "Unknown",
                file_path,
                str(e)
            )
            return False, f"Excel文件损坏或格式错误: {str(e)}"

    def _validate_tenant_isolation(self, file_path: str) -> Tuple[bool, Optional[str]]:
        """
        验证租户隔离（安全加固）

        Args:
            file_path: 文件路径

        Returns:
            (是否有效, 错误消息)
        """
        try:
            # 如果没有租户信息，跳过检查
            if not self.tenant:
                self.logger.debug("没有租户信息，跳过租户隔离检查")
                return True, None

            # 记录租户访问日志
            self.audit_logger.logger.info(
                "租户隔离检查 - 租户: %s (ID: %s), 用户: %s (ID: %s), 文件: %s",
                self.tenant.name if self.tenant else "None",
                self.tenant.id if self.tenant else "None",
                self.user.username if self.user else "None",
                self.user.id if self.user else "None",
                file_path
            )

            # 检查用户是否属于当前租户
            if self.user and hasattr(self.user, 'userprofile'):
                user_tenant = self.user.userprofile.tenant
                if user_tenant != self.tenant:
                    self.logger.error(
                        "租户隔离检查失败：用户租户不匹配 - 用户租户: %s, 当前租户: %s",
                        user_tenant.name if user_tenant else "None",
                        self.tenant.name
                    )
                    self.audit_logger.logger.error(
                        "安全事件：租户隔离违规 - 用户: %s (租户: %s), 尝试访问租户: %s, 文件: %s",
                        self.user.username,
                        user_tenant.name if user_tenant else "None",
                        self.tenant.name,
                        file_path
                    )
                    return False, "租户隔离检查失败：无权访问其他租户的数据"

            # 验证文件路径是否在租户允许的范围内
            # 这里可以根据实际的目录结构进行更严格的检查
            # 例如：检查路径是否包含租户ID或租户名称

            return True, None

        except Exception as e:
            self.logger.error("租户隔离验证失败: %s", str(e), exc_info=True)
            self.audit_logger.logger.error(
                "安全事件：租户隔离验证异常 - 租户: %s, 用户: %s, 文件: %s, 错误: %s",
                self.tenant.name if self.tenant else "None",
                self.user.username if self.user else "None",
                file_path,
                str(e)
            )
            return False, f"租户隔离验证失败: {str(e)}"


    def process_grading_system_scenario(
        self,
        homework_dir: str,
        class_dir: str,
        progress_tracker: Optional[BatchGradeProgressTracker] = None,
    ) -> Dict[str, any]:
        """
        处理作业评分系统场景

        从作业目录扫描Word文档，提取学生姓名和成绩，写入班级目录的登分册

        Args:
            homework_dir: 作业目录路径（如：/path/to/第1次作业）
            class_dir: 班级目录路径（如：/path/to/2024级计算机1班）

        Returns:
            处理结果字典，包含成功、失败、跳过的文件列表和统计信息
        """
        self.logger.info(
            "开始处理作业评分系统场景 - 作业目录: %s, 班级目录: %s",
            homework_dir,
            class_dir,
        )

        # 开始审计记录
        self.audit_logger.start_operation(
            "grading_system_batch_write",
            homework_directory=homework_dir,
            class_directory=class_dir,
        )

        result = {
            "success": False,
            "homework_number": None,
            "registry_path": None,
            "processed_files": [],
            "failed_files": [],
            "skipped_files": [],
            "statistics": {
                "total": 0,
                "success": 0,
                "failed": 0,
                "skipped": 0,
            },
            "error_message": None,
        }

        if progress_tracker:
            progress_tracker.start(message="正在准备批量登分...")

        try:
            # 安全加固：验证作业目录路径
            is_valid, error_msg = self._validate_path_security(homework_dir, homework_dir)
            if not is_valid:
                result["error_message"] = f"作业目录路径验证失败: {error_msg}"
                self.logger.error(result["error_message"])
                if progress_tracker:
                    progress_tracker.fail(result["error_message"])
                return result
            
            # 安全加固：验证班级目录路径
            is_valid, error_msg = self._validate_path_security(class_dir, class_dir)
            if not is_valid:
                result["error_message"] = f"班级目录路径验证失败: {error_msg}"
                self.logger.error(result["error_message"])
                if progress_tracker:
                    progress_tracker.fail(result["error_message"])
                return result
            
            # 安全加固：验证租户隔离
            is_valid, error_msg = self._validate_tenant_isolation(homework_dir)
            if not is_valid:
                result["error_message"] = error_msg
                self.logger.error(result["error_message"])
                if progress_tracker:
                    progress_tracker.fail(result["error_message"])
                return result
            
            # 1. 从作业目录名提取作业批次
            homework_number = GradeFileProcessor.extract_homework_number_from_path(
                homework_dir
            )
            if homework_number is None:
                result["error_message"] = f"无法从目录名提取作业批次: {homework_dir}"
                self.logger.error(result["error_message"])
                if progress_tracker:
                    progress_tracker.fail(result["error_message"])
                return result

            result["homework_number"] = homework_number
            self.logger.info("提取作业批次: %d", homework_number)

            # 2. 查找登分册文件
            registry_path = self.find_grade_registry(class_dir)
            if not registry_path:
                result["error_message"] = f"未找到成绩登分册文件: {class_dir}"
                self.logger.error(result["error_message"])
                if progress_tracker:
                    progress_tracker.fail(result["error_message"])
                return result

            result["registry_path"] = registry_path
            self.logger.info("找到登分册: %s", registry_path)
            
            # 安全加固：验证登分册文件大小
            is_valid, error_msg = self._validate_file_size(registry_path)
            if not is_valid:
                result["error_message"] = f"登分册文件验证失败: {error_msg}"
                self.logger.error(result["error_message"])
                if progress_tracker:
                    progress_tracker.fail(result["error_message"])
                return result
            
            # 安全加固：验证登分册文件完整性
            is_valid, error_msg = self._validate_excel_integrity(registry_path)
            if not is_valid:
                result["error_message"] = f"登分册文件完整性验证失败: {error_msg}"
                self.logger.error(result["error_message"])
                if progress_tracker:
                    progress_tracker.fail(result["error_message"])
                return result

            # 3. 扫描作业目录下的Word文档（性能优化：批量读取文件列表）
            word_files = []
            if not os.path.exists(homework_dir):
                result["error_message"] = f"作业目录不存在: {homework_dir}"
                self.logger.error(result["error_message"])
                if progress_tracker:
                    progress_tracker.fail(result["error_message"])
                return result

            # 性能优化：批量读取文件列表
            for root, _, files in os.walk(homework_dir):
                for file in files:
                    if file.endswith((".docx", ".doc")) and not file.startswith("~$"):
                        word_files.append(os.path.join(root, file))

            if not word_files:
                result["error_message"] = f"作业目录中没有找到Word文档: {homework_dir}"
                self.logger.warning(result["error_message"])
                if progress_tracker:
                    progress_tracker.fail(result["error_message"])
                return result

            result["statistics"]["total"] = len(word_files)
            if progress_tracker:
                progress_tracker.update_total(len(word_files))
            self.logger.info("找到 %d 个Word文档", len(word_files))
            
            # 性能优化：文件数量警告
            if len(word_files) > self.MAX_FILES_WARNING_THRESHOLD:
                self.logger.warning(
                    "文件数量超过阈值 (%d > %d)，处理可能需要较长时间",
                    len(word_files),
                    self.MAX_FILES_WARNING_THRESHOLD
                )
            
            self.logger.info("开始扫描作业目录: %s", homework_dir)
            self.logger.info("识别到 %d 个Word文档文件", len(word_files))

            # 4. 加载登分册
            registry_manager = RegistryManager(registry_path)

            if not registry_manager.load():
                error_message = getattr(registry_manager, "last_error_message", None)
                result["error_message"] = error_message or "加载登分册失败"
                self.logger.error("登分册加载失败: %s", result["error_message"])
                if progress_tracker:
                    progress_tracker.fail(result["error_message"])
                return result

            is_valid, error_msg = registry_manager.validate_format()
            if not is_valid:
                result["error_message"] = error_msg
                if progress_tracker:
                    progress_tracker.fail(result["error_message"])
                return result

            # 5. 创建备份
            if not registry_manager.create_backup():
                result["error_message"] = "创建登分册备份失败"
                if progress_tracker:
                    progress_tracker.fail(result["error_message"])
                return result

            # 6. 处理每个Word文档
            try:
                # 查找或创建作业列
                homework_col = registry_manager.find_or_create_homework_column(
                    homework_number
                )

                for index, word_file in enumerate(word_files, start=1):
                    file_result = self._process_single_word_file(
                        word_file, registry_manager, homework_col
                    )

                    if file_result["success"]:
                        result["processed_files"].append(file_result)
                        result["statistics"]["success"] += 1
                    elif file_result["skipped"]:
                        result["skipped_files"].append(file_result)
                        result["statistics"]["skipped"] += 1
                    else:
                        result["failed_files"].append(file_result)
                        result["statistics"]["failed"] += 1

                    if progress_tracker:
                        progress_tracker.update_progress(
                            processed=index,
                            success=result["statistics"]["success"],
                            failed=result["statistics"]["failed"],
                            skipped=result["statistics"]["skipped"],
                            current_file=os.path.basename(word_file),
                            message=f"正在处理第 {index}/{len(word_files)} 个文件",
                        )

                # 7. 保存登分册
                if registry_manager.save():
                    result["success"] = True
                    registry_manager.delete_backup()
                    self.logger.info(
                        "作业评分系统场景处理完成 - 成功: %d, 失败: %d, 跳过: %d",
                        result["statistics"]["success"],
                        result["statistics"]["failed"],
                        result["statistics"]["skipped"],
                    )

                    # 记录审计日志
                    self.audit_logger.end_operation(
                        success=True,
                        homework_number=homework_number,
                        total_files=result["statistics"]["total"],
                        success_count=result["statistics"]["success"],
                        failed_count=result["statistics"]["failed"],
                        skipped_count=result["statistics"]["skipped"],
                        registry_path=registry_path,
                    )
                    if progress_tracker:
                        progress_tracker.complete(
                            summary=result["statistics"],
                            message="批量登分完成",
                        )
                else:
                    result["error_message"] = "保存登分册失败"
                    registry_manager.restore_from_backup()
                    self.logger.error("保存登分册失败，已从备份恢复")
                    if progress_tracker:
                        progress_tracker.fail(result["error_message"])

                    # 记录审计日志
                    self.audit_logger.end_operation(
                        success=False,
                        error_message=result["error_message"],
                    )

            except Exception as e:
                result["error_message"] = f"处理文件时出错: {str(e)}"
                self.logger.error("处理文件时出错: %s", str(e), exc_info=True)
                registry_manager.restore_from_backup()
                if progress_tracker:
                    progress_tracker.fail(result["error_message"])

                # 记录审计日志
                self.audit_logger.end_operation(
                    success=False,
                    error_message=result["error_message"],
                    exception_type=type(e).__name__,
                )

        except Exception as e:
            result["error_message"] = f"处理作业评分系统场景时出错: {str(e)}"
            self.logger.error(
                "处理作业评分系统场景时出错: %s", str(e), exc_info=True
            )
            if progress_tracker and result["error_message"]:
                progress_tracker.fail(result["error_message"])

            # 记录审计日志
            self.audit_logger.end_operation(
                success=False,
                error_message=result["error_message"],
                exception_type=type(e).__name__,
            )

        if not result["success"] and progress_tracker and result["error_message"]:
            progress_tracker.fail(result["error_message"])

        return result

    def _process_single_word_file(
        self, word_file: str, registry_manager: RegistryManager, homework_col: int
    ) -> Dict[str, any]:
        """
        处理单个Word文档

        Args:
            word_file: Word文档路径
            registry_manager: 登分册管理器
            homework_col: 作业列索引

        Returns:
            处理结果字典
        """
        file_basename = os.path.basename(word_file)

        file_result = {
            "file_path": word_file,
            "file_name": file_basename,
            "student_name": None,
            "grade": None,
            "success": False,
            "skipped": False,
            "error_message": None,
        }

        try:
            # 安全加固：验证文件大小
            is_valid, error_msg = self._validate_file_size(word_file)
            if not is_valid:
                file_result["error_message"] = error_msg
                self.logger.warning("文件大小验证失败: %s - %s", file_basename, error_msg)
                self.audit_logger.log_file_processing(word_file, "failed", error_msg)
                return file_result
            
            # 1. 提取学生姓名
            student_name = GradeFileProcessor.extract_student_name(word_file)
            if not student_name:
                file_result["error_message"] = "无法提取学生姓名"
                self.logger.warning("无法提取学生姓名: %s", file_basename)
                return file_result

            file_result["student_name"] = student_name

            # 2. 验证实验报告评价（需求: 4.5, 5.2, 7.1-7.7）
            is_valid, error_msg = GradeFileProcessor.validate_lab_report_comment(word_file)
            if not is_valid:
                file_result["error_message"] = error_msg
                self.logger.warning("实验报告评价验证失败: %s - %s", file_basename, error_msg)
                self.audit_logger.log_file_processing(word_file, "failed", error_msg)
                return file_result

            # 3. 提取成绩
            grade = GradeFileProcessor.extract_grade_from_word(word_file)
            grade = self._sanitize_grade_value(grade)
            if not grade:
                file_result["error_message"] = "无法提取成绩"
                self.logger.warning("无法提取成绩: %s", file_basename)
                return file_result

            file_result["grade"] = grade

            # 4. 匹配学生姓名
            student_names = list(registry_manager.student_names.keys())
            matched_name, match_type = NameMatcher.match(student_name, student_names)

            if not matched_name:
                filename_match = None
                if match_type != "multiple":
                    filename_match = self._match_student_by_filename(word_file, student_names)
                if filename_match:
                    matched_name = filename_match
                    match_type = "filename"
                    self.logger.info(
                        "通过文件名匹配学生成功: %s -> %s",
                        os.path.basename(word_file),
                        matched_name,
                    )
                else:
                    if match_type == "multiple":
                        file_result["error_message"] = f"姓名匹配到多个学生: {student_name}"
                    else:
                        file_result["error_message"] = f"未找到匹配的学生: {student_name}"
                    self.logger.warning(
                        "%s - 文件: %s", file_result["error_message"], file_basename
                    )
                    return file_result

            # 5. 查找学生行
            student_row = registry_manager.find_student_row(matched_name)
            if not student_row:
                file_result["error_message"] = f"未找到学生行: {matched_name}"
                self.logger.warning(
                    "未找到学生行: %s - 文件: %s", matched_name, file_basename
                )
                return file_result

            # 6. 写入成绩
            written, old_grade = registry_manager.write_grade(
                student_row, homework_col, grade
            )

            if written:
                file_result["success"] = True
                file_result["old_grade"] = old_grade
                self.logger.info(
                    "成功写入成绩 - 学生: %s, 成绩: %s, 文件: %s",
                    matched_name,
                    grade,
                    word_file,
                )

                # 记录审计日志
                homework_number = GradeFileProcessor.extract_homework_number_from_path(word_file)
                if homework_number:
                    self.audit_logger.log_grade_write(
                        matched_name, homework_number, grade, old_grade
                    )

                # 记录文件处理状态
                self.audit_logger.log_file_processing(word_file, "success")
            else:
                file_result["skipped"] = True
                file_result["error_message"] = "成绩相同，跳过写入"
                self.logger.debug(
                    "成绩相同，跳过写入 - 学生: %s, 成绩: %s", matched_name, grade
                )

                # 记录文件处理状态
                self.audit_logger.log_file_processing(
                    word_file, "skipped", "成绩相同，跳过写入"
                )

        except Exception as e:
            file_result["error_message"] = f"处理文件时出错: {str(e)}"
            self.logger.error("处理文件时出错: %s - %s", word_file, str(e), exc_info=True)

            # 记录文件处理状态
            self.audit_logger.log_file_processing(
                word_file, "failed", file_result["error_message"]
            )

        return file_result


    def process_toolbox_scenario(self, class_dir: str) -> Dict[str, any]:
        """
        处理工具箱模块场景

        从班级目录扫描Excel成绩文件，提取作业批次和所有学生成绩，写入登分册

        Args:
            class_dir: 班级目录路径（如：/path/to/2024级计算机1班）

        Returns:
            处理结果字典，包含成功、失败、跳过的文件列表和统计信息
        """
        self.logger.info("开始处理工具箱模块场景 - 班级目录: %s", class_dir)

        # 开始审计记录
        self.audit_logger.start_operation(
            "toolbox_batch_write",
            class_directory=class_dir,
        )

        result = {
            "success": False,
            "registry_path": None,
            "processed_files": [],
            "failed_files": [],
            "skipped_files": [],
            "statistics": {
                "total_files": 0,
                "total_students": 0,
                "success": 0,
                "failed": 0,
                "skipped": 0,
            },
            "error_message": None,
        }

        try:
            # 安全加固：验证班级目录路径
            is_valid, error_msg = self._validate_path_security(class_dir, class_dir)
            if not is_valid:
                result["error_message"] = f"班级目录路径验证失败: {error_msg}"
                self.logger.error(result["error_message"])
                return result
            
            # 安全加固：验证租户隔离
            is_valid, error_msg = self._validate_tenant_isolation(class_dir)
            if not is_valid:
                result["error_message"] = error_msg
                self.logger.error(result["error_message"])
                return result
            
            # 1. 查找登分册文件
            registry_path = self.find_grade_registry(class_dir)
            if not registry_path:
                result["error_message"] = f"未找到成绩登分册文件: {class_dir}"
                self.logger.error(result["error_message"])
                return result

            result["registry_path"] = registry_path
            self.logger.info("找到登分册: %s", registry_path)
            
            # 安全加固：验证登分册文件大小
            is_valid, error_msg = self._validate_file_size(registry_path)
            if not is_valid:
                result["error_message"] = f"登分册文件验证失败: {error_msg}"
                self.logger.error(result["error_message"])
                return result
            
            # 安全加固：验证登分册文件完整性
            is_valid, error_msg = self._validate_excel_integrity(registry_path)
            if not is_valid:
                result["error_message"] = f"登分册文件完整性验证失败: {error_msg}"
                self.logger.error(result["error_message"])
                return result

            # 2. 扫描班级目录下的Excel成绩文件（性能优化：批量读取文件列表）
            excel_files = []
            if not os.path.exists(class_dir):
                result["error_message"] = f"班级目录不存在: {class_dir}"
                self.logger.error(result["error_message"])
                return result

            # 性能优化：批量读取文件列表
            for file in os.listdir(class_dir):
                if file.lower().endswith(".xlsx") and not file.startswith("~$"):
                    # 排除登分册文件本身
                    file_path = os.path.join(class_dir, file)
                    if file_path != registry_path:
                        # 检查文件名是否包含作业批次信息
                        if GradeFileProcessor.extract_homework_number_from_filename(
                            file_path
                        ):
                            excel_files.append(file_path)

            if not excel_files:
                result["error_message"] = f"班级目录中没有找到Excel成绩文件: {class_dir}"
                self.logger.warning(result["error_message"])
                return result

            result["statistics"]["total_files"] = len(excel_files)
            self.logger.info("找到 %d 个Excel成绩文件", len(excel_files))
            
            # 性能优化：文件数量警告
            if len(excel_files) > self.MAX_FILES_WARNING_THRESHOLD:
                self.logger.warning(
                    "文件数量超过阈值 (%d > %d)，处理可能需要较长时间",
                    len(excel_files),
                    self.MAX_FILES_WARNING_THRESHOLD
                )
            
            self.logger.info("开始扫描班级目录: %s", class_dir)
            self.logger.info("识别到 %d 个Excel成绩文件", len(excel_files))

            # 3. 加载登分册
            registry_manager = RegistryManager(registry_path)

            if not registry_manager.load():
                error_message = getattr(registry_manager, "last_error_message", None)
                result["error_message"] = error_message or "加载登分册失败"
                self.logger.error("登分册加载失败: %s", result["error_message"])
                return result

            is_valid, error_msg = registry_manager.validate_format()
            if not is_valid:
                result["error_message"] = error_msg
                return result

            # 4. 创建备份
            if not registry_manager.create_backup():
                result["error_message"] = "创建登分册备份失败"
                return result

            # 5. 处理每个Excel文件
            try:
                for excel_file in excel_files:
                    file_result = self._process_single_excel_file(
                        excel_file, registry_manager
                    )

                    if file_result["success"]:
                        result["processed_files"].append(file_result)
                        result["statistics"]["success"] += file_result[
                            "students_processed"
                        ]
                        result["statistics"]["total_students"] += file_result[
                            "students_total"
                        ]
                    elif file_result["partial_success"]:
                        result["processed_files"].append(file_result)
                        result["statistics"]["success"] += file_result[
                            "students_processed"
                        ]
                        result["statistics"]["failed"] += file_result["students_failed"]
                        result["statistics"]["total_students"] += file_result[
                            "students_total"
                        ]
                    else:
                        result["failed_files"].append(file_result)
                        result["statistics"]["failed"] += 1

                # 6. 保存登分册
                if registry_manager.save():
                    result["success"] = True
                    registry_manager.delete_backup()
                    self.logger.info(
                        "工具箱模块场景处理完成 - 文件: %d, 学生成绩: 成功 %d, 失败 %d",
                        result["statistics"]["total_files"],
                        result["statistics"]["success"],
                        result["statistics"]["failed"],
                    )

                    # 记录审计日志
                    self.audit_logger.end_operation(
                        success=True,
                        total_files=result["statistics"]["total_files"],
                        total_students=result["statistics"]["total_students"],
                        success_count=result["statistics"]["success"],
                        failed_count=result["statistics"]["failed"],
                        registry_path=registry_path,
                    )
                else:
                    result["error_message"] = "保存登分册失败"
                    registry_manager.restore_from_backup()
                    self.logger.error("保存登分册失败，已从备份恢复")

                    # 记录审计日志
                    self.audit_logger.end_operation(
                        success=False,
                        error_message=result["error_message"],
                    )

            except Exception as e:
                result["error_message"] = f"处理文件时出错: {str(e)}"
                self.logger.error("处理文件时出错: %s", str(e), exc_info=True)
                registry_manager.restore_from_backup()

                # 记录审计日志
                self.audit_logger.end_operation(
                    success=False,
                    error_message=result["error_message"],
                    exception_type=type(e).__name__,
                )

        except Exception as e:
            result["error_message"] = f"处理工具箱模块场景时出错: {str(e)}"
            self.logger.error("处理工具箱模块场景时出错: %s", str(e), exc_info=True)

            # 记录审计日志
            self.audit_logger.end_operation(
                success=False,
                error_message=result["error_message"],
                exception_type=type(e).__name__,
            )

        return result

    def _match_student_by_filename(self, word_file: str, student_names: List[str]) -> Optional[str]:
        """如果文件名包含学生姓名，则视为匹配。"""
        filename = os.path.splitext(os.path.basename(word_file))[0]
        normalized_filename = NameMatcher.normalize_name(filename)
        if not normalized_filename:
            return None

        matches = []
        for candidate in student_names:
            normalized_candidate = NameMatcher.normalize_name(candidate)
            if normalized_candidate and normalized_candidate in normalized_filename:
                matches.append(candidate)

        if len(matches) == 1:
            return matches[0]
        if len(matches) > 1:
            self.logger.warning(
                "文件名包含多个学生姓名: %s -> %s", filename, matches
            )
        return None

    @staticmethod
    def _sanitize_grade_value(value) -> Optional[str]:
        """将成绩值转换为非空字符串，过滤 NaN/None。"""
        if value is None:
            return None
        if isinstance(value, float) and math.isnan(value):
            return None
        text = str(value).strip()
        if not text:
            return None
        if text.lower() in {"nan", "none", "null"}:
            return None
        return text

    def _process_single_excel_file(
        self, excel_file: str, registry_manager: RegistryManager
    ) -> Dict[str, any]:
        """
        处理单个Excel成绩文件

        Args:
            excel_file: Excel文件路径
            registry_manager: 登分册管理器

        Returns:
            处理结果字典
        """
        file_result = {
            "file_path": excel_file,
            "homework_number": None,
            "students_total": 0,
            "students_processed": 0,
            "students_failed": 0,
            "success": False,
            "partial_success": False,
            "error_message": None,
            "student_details": [],
        }

        try:
            # 安全加固：验证文件大小
            is_valid, error_msg = self._validate_file_size(excel_file)
            if not is_valid:
                file_result["error_message"] = error_msg
                self.logger.warning("文件大小验证失败: %s - %s", excel_file, error_msg)
                self.audit_logger.log_file_processing(excel_file, "failed", error_msg)
                return file_result
            
            # 安全加固：验证Excel文件完整性
            is_valid, error_msg = self._validate_excel_integrity(excel_file)
            if not is_valid:
                file_result["error_message"] = error_msg
                self.logger.warning("Excel文件完整性验证失败: %s - %s", excel_file, error_msg)
                self.audit_logger.log_file_processing(excel_file, "failed", error_msg)
                return file_result
            
            # 1. 从文件名提取作业批次
            homework_number = GradeFileProcessor.extract_homework_number_from_filename(
                excel_file
            )
            if homework_number is None:
                file_result["error_message"] = "无法从文件名提取作业批次"
                self.logger.warning("无法从文件名提取作业批次: %s", excel_file)
                return file_result

            file_result["homework_number"] = homework_number
            self.logger.info("处理Excel文件 - 作业批次: %d, 文件: %s", homework_number, excel_file)
            self.logger.info("从文件名提取作业批次: %d", homework_number)

            # 2. 从Excel提取所有学生成绩
            grades_data = GradeFileProcessor.extract_grades_from_excel(excel_file)
            if not grades_data:
                file_result["error_message"] = "无法从Excel提取学生成绩"
                self.logger.warning("无法从Excel提取学生成绩: %s", excel_file)
                return file_result

            file_result["students_total"] = len(grades_data)
            self.logger.info("从Excel提取了 %d 个学生成绩", len(grades_data))

            # 3. 查找或创建作业列
            homework_col = registry_manager.find_or_create_homework_column(
                homework_number
            )

            # 4. 处理每个学生成绩（性能优化：缓存学生列表避免重复查询）
            student_names = list(registry_manager.student_names.keys())

            # 性能优化：工具箱场景下批量处理Excel文件中的学生记录
            for grade_data in grades_data:
                student_detail = self._process_single_student_grade(
                    grade_data, registry_manager, homework_col, student_names
                )
                file_result["student_details"].append(student_detail)

                if student_detail["success"]:
                    file_result["students_processed"] += 1
                else:
                    file_result["students_failed"] += 1

            # 5. 判断处理结果
            if file_result["students_failed"] == 0:
                file_result["success"] = True
                self.logger.info(
                    "Excel文件处理成功 - 所有 %d 个学生成绩已写入",
                    file_result["students_processed"],
                )

                # 记录文件处理状态
                self.audit_logger.log_file_processing(excel_file, "success")
            elif file_result["students_processed"] > 0:
                file_result["partial_success"] = True
                self.logger.warning(
                    "Excel文件部分成功 - 成功: %d, 失败: %d",
                    file_result["students_processed"],
                    file_result["students_failed"],
                )

                # 记录文件处理状态
                self.audit_logger.log_file_processing(
                    excel_file,
                    "success",
                    f"部分成功: {file_result['students_processed']}/{file_result['students_total']}",
                )
            else:
                file_result["error_message"] = "所有学生成绩写入失败"
                self.logger.error("Excel文件处理失败 - 所有学生成绩写入失败")

                # 记录文件处理状态
                self.audit_logger.log_file_processing(
                    excel_file, "failed", file_result["error_message"]
                )

        except Exception as e:
            file_result["error_message"] = f"处理Excel文件时出错: {str(e)}"
            self.logger.error(
                "处理Excel文件时出错: %s - %s", excel_file, str(e), exc_info=True
            )

            # 记录文件处理状态
            self.audit_logger.log_file_processing(
                excel_file, "failed", file_result["error_message"]
            )

        return file_result

    def _process_single_student_grade(
        self,
        grade_data: Dict[str, str],
        registry_manager: RegistryManager,
        homework_col: int,
        student_names: List[str],
    ) -> Dict[str, any]:
        """
        处理单个学生成绩

        Args:
            grade_data: 学生成绩数据 {"name": "张三", "grade": "A"}
            registry_manager: 登分册管理器
            homework_col: 作业列索引
            student_names: 登分册中的学生姓名列表

        Returns:
            处理结果字典
        """
        student_detail = {
            "student_name": grade_data["name"],
            "grade": grade_data["grade"],
            "success": False,
            "error_message": None,
        }

        try:
            student_name = grade_data["name"]
            grade = self._sanitize_grade_value(grade_data["grade"])
            if not grade:
                student_detail["error_message"] = "无法提取成绩"
                self.logger.warning("Excel成绩无效: %s", grade_data["name"])
                return student_detail

            # 1. 匹配学生姓名
            matched_name, match_type = NameMatcher.match(student_name, student_names)

            if not matched_name:
                if match_type == "multiple":
                    student_detail["error_message"] = f"姓名匹配到多个学生: {student_name}"
                else:
                    student_detail["error_message"] = f"未找到匹配的学生: {student_name}"
                self.logger.warning(student_detail["error_message"])
                return student_detail

            # 2. 查找学生行
            student_row = registry_manager.find_student_row(matched_name)
            if not student_row:
                student_detail["error_message"] = f"未找到学生行: {matched_name}"
                self.logger.warning(student_detail["error_message"])
                return student_detail

            # 3. 写入成绩
            written, old_grade = registry_manager.write_grade(
                student_row, homework_col, grade
            )

            if written:
                student_detail["success"] = True
                student_detail["old_grade"] = old_grade
                self.logger.debug(
                    "成功写入成绩 - 学生: %s, 成绩: %s", matched_name, grade
                )

                # 记录审计日志 - 从homework_col推断作业批次
                # 注意：这里我们需要从列索引反推作业批次，暂时使用列索引
                self.audit_logger.log_grade_write(
                    matched_name, homework_col, grade, old_grade
                )
            else:
                student_detail["success"] = True  # 跳过也算成功
                student_detail["skipped"] = True
                self.logger.debug("成绩相同，跳过写入 - 学生: %s, 成绩: %s", matched_name, grade)

        except Exception as e:
            student_detail["error_message"] = f"处理学生成绩时出错: {str(e)}"
            self.logger.error(
                "处理学生成绩时出错: %s - %s",
                grade_data["name"],
                str(e),
                exc_info=True,
            )

        return student_detail


    def find_grade_registry(self, class_dir: str) -> Optional[str]:
        """
        查找成绩登分册文件

        在班级目录中查找包含"成绩登分册"或"登分册"的Excel文件

        Args:
            class_dir: 班级目录路径

        Returns:
            登分册文件路径，如果未找到则返回None
        """
        try:
            if not os.path.exists(class_dir):
                self.logger.error("班级目录不存在: %s", class_dir)
                return None

            if not os.path.isdir(class_dir):
                self.logger.error("路径不是目录: %s", class_dir)
                return None

            # 查找登分册文件（要求文件名为“成绩登分册.xlsx”）
            target_name = "成绩登分册.xlsx"
            for file in os.listdir(class_dir):
                if file.startswith("~$"):
                    continue
                if file == target_name:
                    registry_path = os.path.join(class_dir, file)
                    self.logger.info("找到登分册文件: %s", registry_path)
                    return registry_path

            self.logger.warning("未找到登分册文件: %s", class_dir)
            return None

        except Exception as e:
            self.logger.error(
                "查找登分册文件时出错: %s - %s", class_dir, str(e), exc_info=True
            )
            return None

    def process(self, **kwargs) -> Dict[str, any]:
        """
        根据场景处理成绩写入

        Args:
            **kwargs: 场景相关参数
                - 作业评分系统场景: homework_dir, class_dir
                - 工具箱模块场景: class_dir

        Returns:
            处理结果字典
        """
        try:
            if self.scenario == self.SCENARIO_GRADING_SYSTEM:
                homework_dir = kwargs.get("homework_dir")
                class_dir = kwargs.get("class_dir")

                if not homework_dir or not class_dir:
                    return {
                        "success": False,
                        "error_message": "缺少必需参数: homework_dir 和 class_dir",
                    }

                return self.process_grading_system_scenario(homework_dir, class_dir)

            elif self.scenario == self.SCENARIO_TOOLBOX:
                class_dir = kwargs.get("class_dir")

                if not class_dir:
                    return {
                        "success": False,
                        "error_message": "缺少必需参数: class_dir",
                    }

                return self.process_toolbox_scenario(class_dir)

            else:
                return {
                    "success": False,
                    "error_message": f"未知的场景类型: {self.scenario}",
                }

        except Exception as e:
            self.logger.error("处理成绩写入时出错: %s", str(e), exc_info=True)
            return {
                "success": False,
                "error_message": f"处理成绩写入时出错: {str(e)}",
            }
